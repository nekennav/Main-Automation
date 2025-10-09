import streamlit as st
import os
import pandas as pd
import plotly.express as px
import random
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, NamedStyle
from openpyxl.utils import get_column_letter

# Check for xlrd availability
try:
    import xlrd
except ImportError:
    st.error("Missing 'xlrd' library required for .xls files. Install it using: pip install xlrd>=2.0.1")
    st.stop()

# Set page configuration (first Streamlit command)
st.set_page_config(page_title="NYEL", layout="wide", page_icon="💻")

# Initialize session state for page navigation
if 'page' not in st.session_state:
    st.session_state.page = 'home'

# CSS for consistent styling across pages with animated background
st.markdown("""
<style>
/* Apply background image with vertical animation to the Streamlit app container */
.stApp {
    background: linear-gradient(rgba(0, 0, 0, 0.7), rgba(0, 0, 0, 0.7)), url('https://images3.alphacoders.com/134/1342304.jpeg');
    background-size: cover;
    background-position: center;
    background-repeat: no-repeat;
    background-color: #000000; /* Fallback color to prevent white blanks */
    animation: panningBackground 50s linear infinite;
    color: #FFFFFF;
}
@keyframes panningBackground {
    0% { background-position: 50% 0%; }
    50% { background-position: 50% 100%; }
    100% { background-position: 50% 0%; }
}
/* Container with semi-transparent blue background for pages other than MC4 RESHUFFLE */
.container {
    background: rgba(31, 119, 180, 0.85);
    padding: 40px;
    border-radius: 15px;
    box-shadow: 0 8px 16px rgba(0,0,0,0.2);
    margin-bottom: 20px;
    position: relative;
    z-index: 1;
}
/* Enhanced text shadow for all text within container */
.container * {
    text-shadow: -2px -2px 0 #000000, 2px -2px 0 #000000, -2px 2px 0 #000000, 2px 2px 0 #000000,
                 -1px -1px 0 #000000, 1px -1px 0 #000000, -1px 1px 0 #000000, 1px 1px 0 #000000;
    color: #FFFFFF !important;
}
/* Ensure all h1 elements (used by st.title) are white */
h1 {
    color: #FFFFFF !important;
    text-shadow: -2px -2px 0 #000000, 2px -2px 0 #000000, -2px 2px 0 #000000, 2px 2px 0 #000000;
}
/* Back button styling */
.back-button {
    background-color: rgba(31, 119, 180, 0.9);
    border: 2px solid #FFFFFF;
    border-radius: 8px;
    padding: 10px;
    margin-bottom: 20px;
    color: #FFFFFF !important;
    text-shadow: -1px -1px 0 #000000, 1px -1px 0 #000000, -1px 1px 0 #000000, 1px 1px 0 #000000;
}
.back-button:hover {
    background-color: rgba(31, 119, 180, 1.0);
    border-color: #2563EB;
    color: #FFFFFF !important;
}
.main-header {
    font-size: 52px;
    font-weight: bold;
    text-align: center;
    margin-bottom: 20px;
    color: #FFFFFF;
    text-shadow: -2px -2px 0 #000000, 2px -2px 0 #000000, -2px 2px 0 #000000, 2px 2px 0 #000000;
    animation: fadeIn 2s ease-in-out;
}
@keyframes fadeIn {
    0% { opacity: 0; transform: translateY(-20px); }
    100% { opacity: 1; transform: translateY(0); }
}
.sub-header {
    font-size: 28px;
    font-weight: bold;
    color: #FFFFFF;
    text-shadow: -1px -1px 0 #000000, 1px -1px 0 #000000, -1px 1px 0 #000000, 1px 1px 0 #000000;
    margin-top: 30px;
    margin-bottom: 15px;
}
.description {
    font-size: 18px;
    color: #E5E7EB;
    text-align: center;
    margin-bottom: 30px;
    text-shadow: -1px -1px 0 #000000, 1px -1px 0 #000000, -1px 1px 0 #000000, 1px 1px 0 #000000;
}
.feature-card {
    background-color: rgba(255, 255, 255, 0.95);
    padding: 20px;
    border-radius: 10px;
    box-shadow: 0 4px 8px rgba(0,0,0,0.1);
    margin-bottom: 20px;
    transition: transform 0.3s ease;
}
.feature-card:hover {
    transform: translateY(-5px);
    box-shadow: 0 6px 12px rgba(0,0,0,0.15);
}
.icon {
    font-size: 50px;
    margin-bottom: 15px;
    color: #1F77B4;
}
.card-title {
    font-size: 22px;
    font-weight: bold;
    color: #111827;
    margin-bottom: 10px;
}
.card-desc {
    font-size: 16px;
    color: #4B5563;
}
.footer {
    text-align: center;
    font-size: 14px;
    color: #FFFFFF;
    margin-top: 40px;
    padding: 20px;
    background-color: rgba(31, 119, 180, 0.2);
    border-radius: 10px;
    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
}
.stButton > button {
    width: 100%;
    margin-top: 10px;
    background-color: rgba(31, 119, 180, 0.9);
    color: #FFFFFF !important;
    border: 2px solid #FFFFFF;
    border-radius: 8px;
    padding: 10px;
    display: flex;
    align-items: center;
    justify-content: center;
    transition: background-color 0.3s ease, border-color 0.3s ease;
    text-shadow: -1px -1px 0 #000000, 1px -1px 0 #000000, -1px 1px 0 #000000, 1px 1px 0 #000000;
}
.stButton > button:hover {
    background-color: rgba(31, 119, 180, 1.0);
    border-color: #2563EB;
    color: #FFFFFF !important;
}
.stDownloadButton > button {
    background-color: rgba(31, 119, 180, 0.9);
    color: #FFFFFF !important;
    border: 2px solid #FFFFFF;
    border-radius: 8px;
    padding: 10px;
    text-shadow: -1px -1px 0 #000000, 1px -1px 0 #000000, -1px 1px 0 #000000, 1px 1px 0 #000000;
}
.stDownloadButton > button:hover {
    background-color: rgba(31, 119, 180, 1.0);
    border-color: #2563EB;
    color: #FFFFFF !important;
}
.stFileUploader > div > div > label,
.stFileUploader > div > div > span,
.stFileUploader > div > div > div > span,
.stFileUploader > div > div > div > label,
.stFileUploader > div > div > div > div > span,
.stFileUploader > div > div > div > div > label {
    color: #FFFFFF !important;
    text-shadow: 1px 1px 2px rgba(0, 0, 0, 0.5);
}
.stTextArea textarea {
    background-color: rgba(31, 119, 180, 0.2) !important;
    color: #FFFFFF !important;
    border-radius: 8px;
}
.stDataFrame div {
    background-color: rgba(31, 119, 180, 0.2) !important;
}
.stDataFrame table,
.stDataFrame th,
.stDataFrame td {
    color: #FFFFFF !important;
}
.stInfo, .stSuccess, .stWarning, .stError {
    background-color: rgba(31, 119, 180, 0.2) !important;
    color: #FFFFFF !important;
    text-shadow: -1px -1px 0 #000000, 1px -1px 0 #000000, -1px 1px 0 #000000, 1px 1px 0 #000000;
}
.plotly .plotly-graph-div {
    background-color: rgba(31, 119, 180, 0.2) !important;
}
.plotly .plotly-graph-div text,
.plotly .plotly-graph-div tspan,
.plotly .plotly-graph-div .plotly-title,
.plotly .plotly-graph-div .xaxis-title,
.plotly .plotly-graph-div .yaxis-title {
    fill: #FFFFFF !important;
    text-shadow: -1px -1px 0 #000000, 1px -1px 0 #000000, -1px 1px 0 #000000, 1px 1px 0 #000000;
}
.plotly .plotly-graph-div .modebar-btn text {
    fill: #000000 !important;
    text-shadow: none !important;
}
/* Ensure all text in DRR BREAKDOWN and SBC B2 REPORT is white */
h2, h3, p, div, span, .stMetric, .stMetric * {
    color: #FFFFFF !important;
    text-shadow: -1px -1px 0 #000000, 1px -1px 0 #000000, -1px 1px 0 #000000, 1px 1px 0 #000000;
}
/* Ensure all markdown text (used by st.markdown) is white */
.stMarkdown, .stMarkdown *, .stMarkdown p, .stMarkdown div, .stMarkdown span {
    color: #FFFFFF !important;
    text-shadow: -1px -1px 0 #000000, 1px -1px 0 #000000, -1px 1px 0 #000000, 1px 1px 0 #000000;
}
</style>
""", unsafe_allow_html=True)

# Home page
if st.session_state.page == 'home':
    st.markdown('<div class="container">', unsafe_allow_html=True)
    st.markdown('<div class="main-header">"DON\'T BE AFRAID TO FAIL, BE AFRAID NOT TO TRY"</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-header">Our Tools</div>', unsafe_allow_html=True)
    col1, col2 = st.columns([1, 1], gap="medium")
    with col1:
        if st.button("📊 SBC B2 REPORT", key="sbc_b2_button"):
            st.session_state.page = "SBC B2 REPORT"
            st.rerun()
        if st.button("🔍 DRR BREAKDOWN", key="drr_breakdown_button"):
            st.session_state.page = "DRR BREAKDOWN"
            st.rerun()
    with col2:
        if st.button("🔄 MC4 RESHUFFLE", key="mc4_reshuffle_button"):
            st.session_state.page = "MC4 RESHUFFLE"
            st.rerun()
        if st.button("📂 PREDICTIVE MERGER", key="predictive_merger_button"):
            st.session_state.page = "PREDICTIVE MERGER"
            st.rerun()

# SBC B2 REPORT
elif st.session_state.page == "SBC B2 REPORT":
    if st.button("Back to Home", key="back_home_sbc", help="Return to home page", type="secondary"):
        st.session_state.page = 'home'
        st.rerun()
    st.markdown('<div class="container">', unsafe_allow_html=True)
    st.title("SBC B2 REPORT")
    # File uploader widget
    uploaded_files = st.file_uploader(
        "Choose files to upload",
        accept_multiple_files=True,
        type=['txt', 'pdf', 'png', 'jpg', 'jpeg', 'csv', 'xlsx']
    )
    # Create a directory to store uploaded files
    UPLOAD_DIR = "Uploads"
    if not os.path.exists(UPLOAD_DIR):
        os.makedirs(UPLOAD_DIR)
    # List of Manual PTP Remark Types
    MANUAL_PTP_TYPES = ["Follow Up", "SMS", "Email", "Outgoing"]
    # List of PTP Statuses
    PTP_STATUSES = [
        "OUTGOING CALL - PTP NO DISCOUNT",
        "EMAIL BLAST SENT - PTP NO DISCOUNT",
        "FV HAND CARRY - PTP NO DISCOUNT",
        "FIELD VISIT RESULT - PTP NO DISCOUNT",
        "INCOMING CALL - PTP NO DISCOUNT",
        "SMS BLAST SENT - PTP NO DISCOUNT",
        "WITH FIELD RESULT - PTP_NO DISCOUNT"
    ]
    # Function to determine account type based on Account No.
    def get_account_type(account_no):
        if isinstance(account_no, str):
            if account_no.startswith('000'):
                return 'Cards'
            elif account_no.startswith('@BELL'):
                return 'BEL'
        return 'Unknown'
    # Function to count PTPs, manual calls, and get PTP rows for unique accounts
    def get_ptp_counts(df):
        try:
            if 'Remark Type' not in df.columns:
                return {"Error": "No Remark Type column", "PTP Data": None}
            if 'Account No.' not in df.columns:
                return {"Error": "No Account No. column", "PTP Data": None}
            if 'Status' not in df.columns:
                return {"Error": "No Status column", "PTP Data": None}
            
            df['Remark Type'] = df['Remark Type'].str.title()
            df['Status'] = df['Status'].str.title()
            
            df_ptp = df[df['Status'].isin([s.title() for s in PTP_STATUSES])]
            df_ptp = df_ptp[df_ptp['Remark Type'].isin(['Predictive'] + MANUAL_PTP_TYPES)]
            df_unique = df_ptp.drop_duplicates(subset=['Account No.'], keep='first')
            df_unique['Account Type'] = df_unique['Account No.'].apply(get_account_type)
            
            predictive_count = len(df_unique[df_unique['Remark Type'] == 'Predictive'])
            manual_count = len(df_unique[df_unique['Remark Type'].isin(MANUAL_PTP_TYPES)])
            cards_count = len(df_unique[(df_unique['Account Type'] == 'Cards') & (df_unique['Remark Type'] == 'Predictive')])
            bel_count = len(df_unique[(df_unique['Account Type'] == 'BEL') & (df_unique['Remark Type'] == 'Predictive')])
            total_count = len(df_unique)
            
            df_manual_calls = df[
                (df['Remark Type'] == 'Outgoing') &
                (df['Status'].str.contains('OUTGOING CALL -', case=False, na=False))
            ]
            manual_call_unique_count = len(df_manual_calls['Account No.'].drop_duplicates())
            total_manual_calls = len(df_manual_calls)
            avg_manual_calls = total_manual_calls / manual_call_unique_count if manual_call_unique_count > 0 else 0
            
            return {
                'Predictive': predictive_count,
                'Manual': manual_count,
                'Cards': cards_count,
                'BEL': bel_count,
                'Total': total_count,
                'Manual Call Unique Accounts': manual_call_unique_count,
                'Average Manual Calls per Account': avg_manual_calls,
                'PTP Data': df_unique
            }
        except Exception as e:
            return {"Error": f"Error processing data: {str(e)}", "PTP Data": None}
    if uploaded_files:
        st.success(f"Successfully uploaded {len(uploaded_files)} file(s)!")
        for file in uploaded_files:
            file_path = os.path.join(UPLOAD_DIR, file.name)
            with open(file_path, "wb") as f:
                f.write(file.getbuffer())
            st.subheader(f"File: {file.name}")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.write(f"**Size**: {file.size / 1024:.2f} KB")
            with col2:
                st.write(f"**Type**: {file.type}")
            with col3:
                st.write(f"**Uploaded**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            if file.type in ["text/csv", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"]:
                try:
                    if file.type == "text/csv":
                        df = pd.read_csv(file)
                    else:
                        xl = pd.ExcelFile(file, engine='openpyxl')
                        target_sheet = xl.sheet_names[0]
                        st.info(f"Processing sheet: {target_sheet}")
                        df = pd.read_excel(file, engine='openpyxl', sheet_name=target_sheet)
                    ptp_counts = get_ptp_counts(df)
                    if "Error" in ptp_counts:
                        st.warning(ptp_counts["Error"])
                    else:
                        st.subheader("PTP Counts (Unique Accounts)")
                        col1, col2 = st.columns(2)
                        with col1:
                            st.metric("Predictive PTPs", ptp_counts['Predictive'])
                            st.metric("Cards (Predictive)", ptp_counts['Cards'])
                            st.metric("BEL (Predictive)", ptp_counts['BEL'])
                        with col2:
                            st.metric("Manual PTPs", ptp_counts['Manual'])
                            st.metric("Total PTPs", ptp_counts['Total'])
                        st.subheader("Manual Call Metrics")
                        col1, col2 = st.columns(2)
                        with col1:
                            st.metric("Unique Accounts with Manual Calls", ptp_counts['Manual Call Unique Accounts'])
                        with col2:
                            st.metric("Average Manual Calls per Account", f"{ptp_counts['Average Manual Calls per Account']:.2f}")
                        preview_columns = ['Debtor', 'Account No.', 'Status', 'Remark Type', 'Account Type']
                        available_columns = [col for col in preview_columns if col in df.columns or col == 'Account Type']
                        preview_df = ptp_counts['PTP Data'][available_columns].head() if ptp_counts['PTP Data'] is not None else df[available_columns].head()
                        preview_text = preview_df.to_string(index=False) if available_columns else "None of the requested columns found."
                        st.subheader("File Preview")
                        st.text_area(f"Preview of {file.name}", preview_text, height=150, key=f"preview_{file.name}*{hash(file.name)}")
                        if ptp_counts.get('PTP Data') is not None and not ptp_counts['PTP Data'].empty:
                            output = io.BytesIO()
                            if file.type == "text/csv":
                                ptp_counts['PTP Data'].to_csv(output, index=False)
                                mime = "text/csv"
                                ext = ".csv"
                            else:
                                ptp_counts['PTP Data'].to_excel(output, index=False, engine='openpyxl')
                                mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                ext = ".xlsx"
                            output.seek(0)
                            st.download_button(
                                label=f"Download PTP-only {file.name}",
                                data=output,
                                file_name=f"PTP*{file.name.rsplit('.', 1)[0]}{ext}",
                                mime=mime,
                                key=f"download_ptp_{file.name}*{hash(file.name)}"
                            )
                        else:
                            st.info("No PTP records found for download.")
                except Exception as e:
                    st.warning(f"Could not process {file.name}: {str(e)}")
            elif file.type in ["text/plain"]:
                try:
                    content = file.read().decode("utf-8")
                    st.subheader("File Preview")
                    st.text_area(f"Preview of {file.name}", content[:500] + "..." if len(content) > 500 else content, height=150, key=f"preview*{file.name}_{hash(file.name)}")
                    st.info("PTP count not applicable for text files")
                except:
                    st.warning(f"Could not preview {file.name}")
            elif file.type in ["image/png", "image/jpeg", "image/jpg"]:
                st.image(file, caption=f"Preview of {file.name}", use_column_width=True)
                st.info("PTP count not applicable for image files")
            else:
                st.info(f"No preview available for {file.name}")
    st.markdown('</div>', unsafe_allow_html=True)

# DRR BREAKDOWN
elif st.session_state.page == "DRR BREAKDOWN":
    if st.button("Back to Home", key="back_home_drr", help="Return to home page", type="secondary"):
        st.session_state.page = 'home'
        st.rerun()
    st.markdown('<div class="container">', unsafe_allow_html=True)
    st.title("DRR BREAKDOWN")
    uploaded_file = st.file_uploader("Choose an Excel file", type=["xls", "xlsx"], key="file_uploader")
    if uploaded_file is not None:
        try:
            df = pd.read_excel(uploaded_file, engine='xlrd' if uploaded_file.name.endswith('.xls') else 'openpyxl')
            def categorize_source(remark_type):
                if pd.isna(remark_type):
                    return "Unknown"
                remark_type = str(remark_type).lower().strip()
                if 'predictive' in remark_type:
                    return "Predictive"
                elif remark_type in ["sms", "follow up", "email", "outgoing"]:
                    return "SMS" if remark_type == "sms" else "Manual"
                return "Other"
            
            st.subheader("PTP Analysis")
            ptp_statuses = df[
                (df['Status'].str.contains('PTP', case=False, na=False)) &
                (~df['Status'].str.contains('No Commit No PTP', case=False, na=False))
            ]
            if 'Debtor ID' in df.columns and not ptp_statuses.empty:
                ptp_statuses = ptp_statuses.drop_duplicates(subset=['Debtor ID'])
                unique_ptp_count = ptp_statuses['Debtor ID'].nunique()
            else:
                unique_ptp_count = 0
            st.write(f"**Total Unique PTP Accounts**: {unique_ptp_count}")
            if not ptp_statuses.empty and 'Remark Type' in df.columns:
                ptp_statuses['Source'] = ptp_statuses['Remark Type'].apply(categorize_source)
                ptp_source_counts = ptp_statuses.groupby('Source')['Debtor ID'].nunique().reset_index()
                ptp_source_counts.columns = ['Source', 'Unique PTP Count']
                ptp_source_counts = ptp_source_counts.sort_values(by='Unique PTP Count', ascending=False)
                st.write("**PTP Breakdown by Source**:")
                for _, row in ptp_source_counts.iterrows():
                    st.write(f"- {row['Source']}: {row['Unique PTP Count']} accounts")
                fig_ptp = px.bar(
                    ptp_source_counts,
                    x='Source',
                    y='Unique PTP Count',
                    title='Unique PTP Accounts by Source',
                    labels={'Source': 'Source', 'Unique PTP Count': 'Number of Unique PTP Accounts'},
                    color='Source',
                    color_discrete_map={'Predictive': '#FF4B4B', 'Manual': '#1F77B4', 'SMS': '#2CA02C', 'Other': '#7F7F7F', 'Unknown': '#D62728'},
                    text='Unique PTP Count'
                )
                fig_ptp.update_traces(textposition='inside', textfont=dict(size=12, color='white'))
                fig_ptp.update_layout(showlegend=False, xaxis_title="Source", yaxis_title="Unique PTP Accounts", xaxis={'categoryorder': 'array', 'categoryarray': ptp_source_counts['Source'].tolist()}, paper_bgcolor='rgba(31,119,180,0.2)', plot_bgcolor='rgba(31,119,180,0.2)', font_color='#FFFFFF')
                st.plotly_chart(fig_ptp, use_container_width=True)
            else:
                st.write("No PTP records or 'Remark Type' column missing.")
            
            st.subheader("RPC Analysis")
            rpc_statuses_list = [
                'OUTBOUND CALLING - REFUSE TO PAY', 'OUTBOUND CALLING - RETURN CALL',
                'OUTBOUND CALLING - REQUESTING FOR RPS', 'OUTBOUND CALLING - DECEASED',
                'JUNK - NIOP', 'JUNK - DECEASED', 'Positive - LEAVE MESSAGE',
                'Positive Contact - UNDERNEGO', 'Positive Contact - TFIP',
                'Positive Contact - NO INTENTION OF PAYING', 'INBOUND CALLS - COMPLAINT',
                'OUTBOUND CALLING - COMPLAINT', 'OUTBOUND CALLING - NO COMMIT',
                'OUTBOUND CALLING - CANNOT PAY', 'EMAIL - COMPLAINT', 'FIELD - COMPLAINT',
                'INBOUND CALLS - REQUESTING FOR RPS', 'OUTBOUND CALLING - SOA_HOP REQUEST',
                'INBOUND CALLS - RETURN CALL',
                'EMAIL BLAST SENT - NO COMMIT NO PTP', 'EMAIL BLAST SENT - UNDER NEGO',
                'FOLLOW UP CALL - LEFT MSG TO RETURN CALL', 'FV HAND CARRY - LEFT MSG TO RETURN CALL',
                'FV HAND CARRY - NO COMMIT NO PTP', 'FV HAND CARRY - UNDER NEGO',
                'FV HAND CARRY - WITH PENDING REQUEST', 'INCOMING CALL - DECEASED',
                'INCOMING CALL - LEFT MSG TO RETURN CALL', 'INCOMING CALL - NO COMMIT NO PTP',
                'INCOMING CALL - UNDER NEGO', 'INCOMING CALL - WITH PENDING REQUEST',
                'OUTGOING CALL - DECEASED', 'OUTGOING CALL - LEFT MSG TO RETURN CALL',
                'OUTGOING CALL - NO COMMIT NO PTP', 'OUTGOING CALL - UNDER NEGO',
                'OUTGOING CALL - WITH PENDING REQUEST', 'SMS BLAST SENT - LEFT MESSAGE TO RETURN CALL',
                'SMS BLAST SENT - UNDER NEGO', 'VIBER BLAST SENT - UNDER NEGO',
                'WITH FIELD RESULT - DECEASED', 'WITH FIELD RESULT - RPC NIOP_FOR LEGAL ENDO',
                'WITH FIELD RESULT - UNDER NEGO'
            ]
            rpc_statuses = df[df['Status'].isin(rpc_statuses_list)]
            if 'Debtor ID' in df.columns and not rpc_statuses.empty:
                rpc_statuses = rpc_statuses.drop_duplicates(subset=['Debtor ID'])
                unique_rpc_count = rpc_statuses['Debtor ID'].nunique()
            else:
                unique_rpc_count = 0
            st.write(f"**Total Unique RPC Accounts**: {unique_rpc_count}")
            if not rpc_statuses.empty and 'Remark Type' in df.columns:
                rpc_statuses['Source'] = rpc_statuses['Remark Type'].apply(categorize_source)
                rpc_source_counts = rpc_statuses.groupby('Source')['Debtor ID'].nunique().reset_index()
                rpc_source_counts.columns = ['Source', 'Unique RPC Count']
                rpc_source_counts = rpc_source_counts.sort_values(by='Unique RPC Count', ascending=False)
                st.write("**RPC Breakdown by Source**:")
                for _, row in rpc_source_counts.iterrows():
                    st.write(f"- {row['Source']}: {row['Unique RPC Count']} accounts")
                fig_rpc = px.bar(
                    rpc_source_counts,
                    x='Source',
                    y='Unique RPC Count',
                    title='Unique RPC Accounts by Source',
                    labels={'Source': 'Source', 'Unique RPC Count': 'Number of Unique RPC Accounts'},
                    color='Source',
                    color_discrete_map={'Predictive': '#FF4B4B', 'Manual': '#1F77B4', 'SMS': '#2CA02C', 'Other': '#7F7F7F', 'Unknown': '#D62728'},
                    text='Unique RPC Count'
                )
                fig_rpc.update_traces(textposition='inside', textfont=dict(size=12, color='white'))
                fig_rpc.update_layout(showlegend=False, xaxis_title="Source", yaxis_title="Unique RPC Accounts", xaxis={'categoryorder': 'array', 'categoryarray': rpc_source_counts['Source'].tolist()}, paper_bgcolor='rgba(31,119,180,0.2)', plot_bgcolor='rgba(31,119,180,0.2)', font_color='#FFFFFF')
                st.plotly_chart(fig_rpc, use_container_width=True)
            else:
                st.write("No RPC records or 'Remark Type' column missing.")
            
            st.subheader("Claiming Paid Analysis")
            claiming_paid_statuses_list = [
                'INBOUND CALLS - CONFIRMED PAYMENT', 'NEGATIVE - CALL_MOVED OUT_CONFIRMED CLIENT NLR',
                'SMS - CONFIRMED PAYMENT', 'EMAIL - CONFIRMED PAYMENT', 'OUTBOUND CALLING - CONFIRMED PAYMENT',
                'EMAIL - CONFIRMED RPS', 'SMS - CONFIRMED RPS', 'OUTBOUND CALLING - CONFIRMED PARTIAL',
                'EMAIL - CONFIRMED SPIFF OTP', 'EMAIL - CONFIRMED SPIFF INSTALLMENT', 'SMS - CONFIRMED SPIFF OTP',
                'OUTBOUND CALLING - CONFIRMED RPS', 'SMS - CONFIRMED PERENNIAL', 'SMS - CONFIRMED SPIFF INSTALLMENT',
                'OUTBOUND CALLING - CONFIRMED SPIFFINSTALLMENT', 'EMAIL - CONFIRMED PERENNIAL',
                'FIELD - CONFIRMED RPS', 'OUTBOUND CALLING - CONFIRMED SPIFF OTP', 'EMAIL - CONFIRMED OTP',
                'INBOUND CALLS - CONFIRMED SPIFF OTP', 'FIELD - CONFIRMED PARTIAL', 'FIELD - CONFIRMED SPIFF OTP',
                'SMS - CONFIRMED PARTIAL', 'EMAIL - CONFIRMED PARTIAL',
                'CEASE COLL EFFORT SBC - CLAIMING PAID', 'EMAIL BLAST SENT - CLAIMING PAID',
                'FOLLOW UP CALL - CLAIMING PAID', 'FV HAND CARRY - CLAIMING PAID',
                'INCOMING CALL - CLAIMING PAID', 'OUTGOING CALL - CLAIMING PAID',
                'PAYMENT - FULL UPDATE', 'PAYMENT - FULLY PAID', 'PAYMENT - REPO',
                'SMS BLAST SENT - CLAIMING PAID', 'WITH FIELD RESULT - CLAIMING PAID'
            ]
            claiming_paid_statuses = df[df['Status'].isin(claiming_paid_statuses_list)]
            if 'Debtor ID' in df.columns and not claiming_paid_statuses.empty:
                claiming_paid_statuses = claiming_paid_statuses.drop_duplicates(subset=['Debtor ID'])
                unique_claiming_paid_count = claiming_paid_statuses['Debtor ID'].nunique()
            else:
                unique_claiming_paid_count = 0
            st.write(f"**Total Unique Claiming Paid Accounts**: {unique_claiming_paid_count}")
            if not claiming_paid_statuses.empty and 'Remark Type' in df.columns:
                claiming_paid_statuses['Source'] = claiming_paid_statuses['Remark Type'].apply(categorize_source)
                claiming_paid_source_counts = claiming_paid_statuses.groupby('Source')['Debtor ID'].nunique().reset_index()
                claiming_paid_source_counts.columns = ['Source', 'Unique Claiming Paid Count']
                claiming_paid_source_counts = claiming_paid_source_counts.sort_values(by='Unique Claiming Paid Count', ascending=False)
                st.write("**Claiming Paid Breakdown by Source**:")
                for _, row in claiming_paid_source_counts.iterrows():
                    st.write(f"- {row['Source']}: {row['Unique Claiming Paid Count']} accounts")
                fig_claiming_paid = px.bar(
                    claiming_paid_source_counts,
                    x='Source',
                    y='Unique Claiming Paid Count',
                    title='Unique Claiming Paid Accounts by Source',
                    labels={'Source': 'Source', 'Unique Claiming Paid Count': 'Number of Unique Claiming Paid Accounts'},
                    color='Source',
                    color_discrete_map={'Predictive': '#FF4B4B', 'Manual': '#1F77B4', 'SMS': '#2CA02C', 'Other': '#7F7F7F', 'Unknown': '#D62728'},
                    text='Unique Claiming Paid Count'
                )
                fig_claiming_paid.update_traces(textposition='inside', textfont=dict(size=12, color='white'))
                fig_claiming_paid.update_layout(showlegend=False, xaxis_title="Source", yaxis_title="Unique Claiming Paid Accounts", xaxis={'categoryorder': 'array', 'categoryarray': claiming_paid_source_counts['Source'].tolist()}, paper_bgcolor='rgba(31,119,180,0.2)', plot_bgcolor='rgba(31,119,180,0.2)', font_color='#FFFFFF')
                st.plotly_chart(fig_claiming_paid, use_container_width=True)
            else:
                st.write("No Claiming Paid records or 'Remark Type' column missing.")
            
            st.subheader("Summary of Totals")
            st.write(f"**Total Unique PTP Accounts**: {unique_ptp_count}")
            st.write(f"**Total Unique RPC Accounts**: {unique_rpc_count}")
            st.write(f"**Total Unique Claiming Paid Accounts**: {unique_claiming_paid_count}")
        except Exception as e:
            st.error(f"Error reading the file: {str(e)}")
    st.markdown('</div>', unsafe_allow_html=True)

# MC4 RESHUFFLE
elif st.session_state.page == "MC4 RESHUFFLE":
    if st.button("Back to Home", key="back_home_mc4", help="Return to home page", type="secondary"):
        st.session_state.page = 'home'
        st.rerun()
    st.title("MC4 RESHUFFLE")
    @st.cache_resource
    def load_accounts(file):
        try:
            df = pd.read_excel(file)
            required_columns = ['Debtor ID', 'Batch No.', 'Name', 'Account No.', 'Cycle']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                st.error(f"Excel file must contain these columns: {', '.join(missing_columns)}")
                return None, None
            if 'Collector' not in df.columns:
                df['Collector'] = ''
            # Load workbook to preserve formatting
            wb = pd.ExcelFile(file, engine='openpyxl').book
            return df, wb
        except Exception as e:
            st.error(f"Error reading Excel file: {e}")
            return None, None
    def get_collectors(batch_numbers, df):
        batch_numbers_str = ' '.join(batch_numbers.astype(str).str.upper())
        cycle_collectors = {
            (5,): ["KMCRISOSTOMO", "BNOSIA", "RJRAZON", "MADANTAYANA", "NSINADJAN"],
            (12,): ["LAATON", "CHCALFOFORO", "JABIOG", "LEALCANTARA", "RMELENDEZ"],
            (17,): ["ADSARMIENTO", "RMGALSIM", "MGMADAYAG", "JBASOY"],
            (24,): ["NVMAMIGO", "LCSERVALLOS", "JELGARCIA", "JDDAGANIO"],
            (2, 9, 14, 20, 27): ["KAPILAPIL", "EECRUZ", "MGARBAS"]
        }
        if 'SALAD' in batch_numbers_str:
            return ["EHFRANCIA", "JARELUCIO", "JEGUADALUPE", "DAATON", "RTABION", "SNAZURES", "KMHORCA", "RLCORPUZ", "DPVENIEGAS", "JDAMPONG"], "SBF_SALAD", None, []
        elif 'SBF_PL' in batch_numbers_str or 'SBF_LEGACY' in batch_numbers_str:
            return ["RCBANICO", "JBDECHAVEZ", "IMMUNOZ", "BCBAGAYAS", "JEFERRER", "JCANCINO", "VGPARIS", "JBRESULTAY", "MGDIZON", "MCSOLIS", "SARODRIGUEZ", "ECAMADO", "MCMACATIGBAC", "LEPALCE", "JQGAGAM", "ERDEGUZMAN"], "SBF_PL", None, []
        elif 'SBC_B4' in batch_numbers_str:
            return ["PCLAGARIO", "JVVINCULADO", "NBSALIGUMBA", "VMGORDON", "RCFANUNCIANO", "CPPERFAS", "BCBATAC"], "SBC_B4", None, []
        elif 'SBC_B2' in batch_numbers_str:
            # Combine all collectors for SBC_B2 to display in preview
            all_collectors = list(set(
                cycle_collectors[(5,)] +
                cycle_collectors[(12,)] +
                cycle_collectors[(17,)] +
                cycle_collectors[(24,)] +
                cycle_collectors[(2, 9, 14, 20, 27)]
            ))
            # Map cycles to collectors for assignment
            cycle_map = {}
            for cycle_group, collectors in cycle_collectors.items():
                for cycle in cycle_group:
                    cycle_map[cycle] = collectors
            return cycle_map, "SBC_B2", None, all_collectors
        return {}, None, None, []
    def reshuffle_collectors(accounts, cycle_map_or_collectors, campaign):
        shuffled = accounts.copy()
        if campaign == "SBC_B2":
            cycle_map = cycle_map_or_collectors
            # Get unique cycles in the data
            all_cycles = set(shuffled['Cycle'].dropna().astype(int).unique())
            specified_cycles = set(cycle_map.keys())
            valid_cycles = all_cycles.intersection(specified_cycles)
            if not valid_cycles:
                st.error(f"No valid cycles found in the uploaded file. Valid cycles are: {', '.join(map(str, sorted(specified_cycles)))}")
                return shuffled
            for cycle in valid_cycles:
                collectors = cycle_map.get(cycle, [])
                if not collectors:
                    st.warning(f"No collectors defined for Cycle {cycle}. Skipping reshuffle for this cycle.")
                    continue
                cycle_accounts = shuffled[shuffled['Cycle'] == cycle]
                if cycle_accounts.empty:
                    st.warning(f"No accounts found for Cycle {cycle}.")
                    continue
                num_accounts = len(cycle_accounts)
                num_collectors = len(collectors)
                if num_accounts > 0 and num_collectors > 0:
                    # Calculate equal distribution: base + extras to first few collectors
                    base_accounts = num_accounts // num_collectors
                    remainder = num_accounts % num_collectors
                    # Create assignment counts for each collector
                    assignment_counts = [base_accounts + 1 for _ in range(remainder)] + [base_accounts for _ in range(num_collectors - remainder)]
                    
                    # Create the assignments list with exact counts
                    assignments = []
                    for i, collector in enumerate(collectors):
                        assignments.extend([collector] * assignment_counts[i])
                    
                    # Shuffle the assignments to randomize
                    random.shuffle(assignments)
                    
                    # Get the indices of cycle accounts and shuffle them too for fair distribution
                    account_indices = list(cycle_accounts.index)
                    random.shuffle(account_indices)
                    
                    # Now, to avoid assigning the same collector if possible, we'll try to match
                    # but preserve the exact counts per collector
                    original_collectors = cycle_accounts['Collector'].copy()
                    
                    # Create a mapping of available slots per collector
                    collector_slots = {collector: list(range(len(assignments))) for collector in collectors}
                    for i, assignment in enumerate(assignments):
                        collector_slots[assignment].append(i)
                    
                    assigned_slots = {collector: [] for collector in collectors}
                    
                    # Assign accounts to collectors while trying to avoid original
                    for account_idx in account_indices:
                        original_collector = original_collectors.loc[account_idx]
                        
                        # Find a suitable collector (prefer not original, and one with remaining slots)
                        possible_collectors = []
                        for collector in collectors:
                            if collector != original_collector and len(assigned_slots[collector]) < assignment_counts[collectors.index(collector)]:
                                possible_collectors.append(collector)
                        
                        if not possible_collectors:
                            # If no alternatives, use any collector with slots
                            possible_collectors = [c for c in collectors if len(assigned_slots[c]) < assignment_counts[collectors.index(c)]]
                        
                        if possible_collectors:
                            # Choose the one with fewest assignments so far (for balance, though counts are fixed)
                            selected_collector = min(possible_collectors, key=lambda c: len(assigned_slots[c]))
                            # Assign
                            slot_index = collector_slots[selected_collector].pop(0)
                            assignments[slot_index] = selected_collector
                            assigned_slots[selected_collector].append(account_idx)
                            shuffled.at[account_idx, 'Collector'] = selected_collector
                        else:
                            # Fallback: assign arbitrarily while maintaining counts
                            for collector in collectors:
                                if len(assigned_slots[collector]) < assignment_counts[collectors.index(collector)]:
                                    slot_index = collector_slots[collector].pop(0)
                                    assignments[slot_index] = collector
                                    assigned_slots[collector].append(account_idx)
                                    shuffled.at[account_idx, 'Collector'] = collector
                                    break
        else:
            collectors = cycle_map_or_collectors
            if not collectors:
                return shuffled
            num_accounts = len(shuffled)
            num_collectors = len(collectors)
            if num_accounts > 0 and num_collectors > 0:
                # Calculate equal distribution for non-SBC_B2 campaigns
                base_accounts = num_accounts // num_collectors
                remainder = num_accounts % num_collectors
                assignment_counts = [base_accounts + 1 for _ in range(remainder)] + [base_accounts for _ in range(num_collectors - remainder)]
                
                # Create assignments
                assignments = []
                for i, collector in enumerate(collectors):
                    assignments.extend([collector] * assignment_counts[i])
                
                random.shuffle(assignments)
                
                # Shuffle indices
                account_indices = list(shuffled.index)
                random.shuffle(account_indices)
                
                # Avoid original where possible, preserving counts
                original_collectors = shuffled['Collector'].copy()
                collector_slots = {collector: list(range(len(assignments))) for collector in collectors}
                assigned_slots = {collector: [] for collector in collectors}
                
                for account_idx in account_indices:
                    original_collector = original_collectors.loc[account_idx]
                    
                    possible_collectors = []
                    for collector in collectors:
                        if collector != original_collector and len(assigned_slots[collector]) < assignment_counts[collectors.index(collector)]:
                            possible_collectors.append(collector)
                    
                    if not possible_collectors:
                        possible_collectors = [c for c in collectors if len(assigned_slots[c]) < assignment_counts[collectors.index(c)]]
                    
                    if possible_collectors:
                        selected_collector = min(possible_collectors, key=lambda c: len(assigned_slots[c]))
                        slot_index = collector_slots[selected_collector].pop(0)
                        assignments[slot_index] = selected_collector
                        assigned_slots[selected_collector].append(account_idx)
                        shuffled.at[account_idx, 'Collector'] = selected_collector
                    else:
                        # Fallback
                        for collector in collectors:
                            if len(assigned_slots[collector]) < assignment_counts[collectors.index(collector)]:
                                slot_index = collector_slots[collector].pop(0)
                                assignments[slot_index] = collector
                                assigned_slots[collector].append(account_idx)
                                shuffled.at[account_idx, 'Collector'] = collector
                                break
        return shuffled
    def apply_excel_formatting(writer, df, wb):
        worksheet = writer.sheets['Sheet1']
        # Apply header formatting
        header_fill = PatternFill(start_color='1F77B4', end_color='1F77B4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for col_idx, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # Copy column widths from original workbook
        if wb and 'Sheet1' in wb.sheetnames:
            ws = wb['Sheet1']
            for col_idx, column in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_idx)
                try:
                    worksheet.column_dimensions[col_letter].width = ws.column_dimensions[col_letter].width or 15
                except:
                    worksheet.column_dimensions[col_letter].width = 15
        # Apply text style to string columns
        text_style = NamedStyle(name='text', number_format='@')
        if 'text' not in writer.book.named_styles:
            writer.book.add_named_style(text_style)
        for col_idx, column in enumerate(df.columns, 1):
            if df[column].dtype == 'object':
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.style = 'text'
    def main():
        uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx", "xls"])
        accounts_df = None
        original_wb = None
        collectors = []
        campaign = None
        all_collectors = []
        if uploaded_file is not None:
            accounts_df, original_wb = load_accounts(uploaded_file)
            if accounts_df is not None:
                cycle_map, campaign, _, all_collectors = get_collectors(accounts_df['Batch No.'], accounts_df)
                if campaign == "SBC_B2" and all_collectors:
                    st.write(f"**Campaign: {campaign}**")
                    st.write(f"Collectors Assigned for SBC_B2: {', '.join(sorted(all_collectors))}")
                elif cycle_map or collectors:
                    collectors = cycle_map
                    st.write(f"**Campaign: {campaign}**")
                    st.write(f"Collectors Assigned: {', '.join(sorted(collectors))}")
                else:
                    st.error("No collectors available. Ensure Batch No. contains 'SBF_SALAD', 'SBF_PL', 'SBC_B4', or 'SBC_B2' with a valid cycle.")
        if st.button("Reshuffle Collectors"):
            if accounts_df is None:
                st.error("Please upload a valid Excel file with 'Debtor ID', 'Name', 'Batch No.', 'Account No.', and 'Cycle' columns.")
                return
            elif not (cycle_map or collectors) or not campaign:
                st.error("No collectors available. Ensure Batch No. contains 'SBF_SALAD', 'SBF_PL', 'SBC_B4', or 'SBC_B2' and a valid cycle is present.")
                return
            else:
                result_df = reshuffle_collectors(accounts_df, cycle_map if campaign == "SBC_B2" else collectors, campaign)
                st.subheader(f"Reshuffled Account Assignments for {campaign}")
                st.dataframe(result_df, use_container_width=True, hide_index=True)
                current_date = datetime.now().strftime("%m%d%y")
                campaign_filename = f"{campaign.replace(' ', '*')}_{current_date}_RESHUFFLE.xlsx"
                output_buffer = io.BytesIO()
                with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
                    result_df.to_excel(writer, index=False, sheet_name='Sheet1')
                    apply_excel_formatting(writer, result_df, original_wb)
                output_buffer.seek(0)
                st.download_button(
                    label="Download Reshuffled Assignments",
                    data=output_buffer,
                    file_name=campaign_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    if __name__ == "__main__":
        main()

# PREDICTIVE MERGER
elif st.session_state.page == "PREDICTIVE MERGER":
    if st.button("Back to Home", key="back_home_predictive", help="Return to home page", type="secondary"):
        st.session_state.page = 'home'
        st.rerun()
    st.markdown('<div class="container">', unsafe_allow_html=True)
    st.title("PREDICTIVE MERGER")
    uploaded_files = st.file_uploader(
        "Choose Excel files to merge",
        type=["xls", "xlsx"],
        accept_multiple_files=True
    )
    if uploaded_files:
        try:
            sheets_by_name = {}
            preview_dataframes = []
            for uploaded_file in uploaded_files:
                try:
                    excel_data = pd.read_excel(uploaded_file, sheet_name=None, engine='xlrd' if uploaded_file.name.endswith('.xls') else 'openpyxl', dtype={'Phone Number': str})
                    for sheet_name, df in excel_data.items():
                        safe_sheet_name = sheet_name[:31]
                        if safe_sheet_name in sheets_by_name:
                            sheets_by_name[safe_sheet_name].append(df)
                        else:
                            sheets_by_name[safe_sheet_name] = [df]
                except Exception as e:
                    st.error(f"Error processing {uploaded_file.name}: {str(e)}")
                    continue
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"merged_excel_{timestamp}.xlsx"
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                text_style = NamedStyle(name='text', number_format='@')
                if 'text' not in writer.book.named_styles:
                    writer.book.add_named_style(text_style)
                for sheet_name, df_list in sheets_by_name.items():
                    try:
                        merged_df = pd.concat(df_list, ignore_index=True)
                        if 'Phone Number' in merged_df.columns:
                            merged_df['Phone Number'] = merged_df['Phone Number'].astype(str).str.replace('.0$', '', regex=True)
                        if sheet_name == 'Call Status Summary':
                            if 'Call Status' in merged_df.columns:
                                merged_df['Call Status'] = merged_df['Call Status'].fillna('NA')
                                numeric_cols = merged_df.select_dtypes(include=['int64', 'float64']).columns
                                numeric_cols = [col for col in numeric_cols if col != 'Phone Number']
                                agg_dict = {col: 'sum' for col in numeric_cols}
                                if 'Description' in merged_df.columns:
                                    agg_dict['Description'] = 'first'
                                if agg_dict:
                                    aggregated_df = merged_df.groupby('Call Status').agg(agg_dict).reset_index()
                                    desired_columns = ['Call Status']
                                    if 'Description' in aggregated_df.columns:
                                        desired_columns.append('Description')
                                    if 'Count' in aggregated_df.columns:
                                        desired_columns.append('Count')
                                    remaining_columns = [col for col in aggregated_df.columns if col not in desired_columns]
                                    final_columns = desired_columns + remaining_columns
                                    merged_df = aggregated_df[final_columns]
                        merged_df.to_excel(writer, index=False, sheet_name=sheet_name)
                        if 'Phone Number' in merged_df.columns:
                            worksheet = writer.sheets[sheet_name]
                            phone_col_idx = merged_df.columns.get_loc('Phone Number') + 1
                            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=phone_col_idx, max_col=phone_col_idx):
                                for cell in row:
                                    cell.style = 'text'
                        preview_dataframes.append((sheet_name, merged_df))
                    except Exception as e:
                        st.error(f"Error merging sheet {sheet_name}: {str(e)}")
                        continue
            if preview_dataframes:
                try:
                    preview_df = pd.concat([df for _, df in preview_dataframes], ignore_index=True)
                    st.write("**Preview of Data in Merged File:**")
                    st.dataframe(preview_df.head(), use_container_width=True)
                except Exception as e:
                    st.error(f"Error generating preview: {str(e)}")
            output.seek(0)
            st.download_button(
                label="Download Merged Excel File",
                data=output,
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Error creating merged file: {str(e)}")
    st.markdown('</div>', unsafe_allow_html=True)

